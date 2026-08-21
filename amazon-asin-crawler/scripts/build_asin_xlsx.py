#!/usr/bin/env python3
"""Convert the crawler JSON output into 亚马逊前3页ASIN列表.xlsx."""

import json
from datetime import date
import pathlib
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


BASE_HEADERS = ["ASIN", "产品标题", "产品价格", "星级评分", "评论数量", "产品首图链接"]
BASE_WIDTHS = [14, 60, 14, 12, 12, 60]


def dated_output_path(base: pathlib.Path, stem: str, extension: str) -> pathlib.Path:
    """Return base/<stem>-YYYYMMDD.ext, adding -01/-02 when already present."""
    stamp = date.today().strftime("%Y%m%d")
    candidate = base / f"{stem}-{stamp}.{extension}"
    index = 1
    while candidate.exists() and index <= 99:
        candidate = base / f"{stem}-{stamp}-{index:02d}.{extension}"
        index += 1
    return candidate


def avoid_overwrite(path: pathlib.Path) -> pathlib.Path:
    """Append -01/-02 to an explicit path when it already exists."""
    if not path.exists():
        return path
    for index in range(1, 100):
        candidate = path.with_name(
            f"{path.stem}-{index:02d}{path.suffix}"
        )
        if not candidate.exists():
            return candidate
    return path


def main() -> int:
    base = pathlib.Path(__file__).resolve().parent
    json_path = (
        pathlib.Path(sys.argv[1]).resolve()
        if len(sys.argv) > 1
        else base / "asin_raw_results.json"
    )
    output_path = dated_output_path(base, "亚马逊前3页ASIN列表", "xlsx")
    if len(sys.argv) > 2:
        output_path = avoid_overwrite(pathlib.Path(sys.argv[2]).resolve())

    data = json.loads(json_path.read_text(encoding="utf-8"))
    if data.get("status") not in (None, "ok"):
        print(f"crawler status is {data.get('status')!r}; not building xlsx")
        return 1

    rows = data.get("rows") or []
    headers = (
        ["关键词"] + BASE_HEADERS
        if any("关键词" in (row or {}) for row in rows)
        else BASE_HEADERS
    )
    widths = [18] + BASE_WIDTHS if headers[0] == "关键词" else BASE_WIDTHS
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "ASIN列表"

    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center")

    for row in rows:
        sheet.append([row.get(header, "") for header in headers])

    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A2"

    workbook.save(output_path)
    print(f"OK rows={len(rows)} output={output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
